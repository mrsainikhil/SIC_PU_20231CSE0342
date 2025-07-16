import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create the data
data = {
    "EV Model": ["Ola S1 Pro", "TVS iQube", "Ather 450X", "Bajaj Chetak", "Hero Vida V1"],
    "Battery Capacity (kWh)": [3.97, 3.4, 3.7, 3.0, 3.44],
    "Replacement Cost Min (₹)": [45000, 40000, 50000, 35000, 40000],
    "Replacement Cost Max (₹)": [60000, 55000, 65000, 50000, 55000],
    "Lifespan Years Min": [5, 6, 6, 5, 5],
    "Lifespan Years Max": [8, 8, 8, 7, 7],
    "Lifespan KM": [80000, 70000, 80000, 60000, 70000],
    "Warranty Years": ["3-5", "3-5", "3-5", "3-5", "3-5"],
    "Warranty KM": ["60k-80k", "60k-80k", "60k-80k", "60k-80k", "60k-80k"]
}

# Create DataFrame
df = pd.DataFrame(data)

# Calculate additional analysis columns
df["Replacement Cost Avg (₹)"] = (df["Replacement Cost Min (₹)"] + df["Replacement Cost Max (₹)"]) / 2
df["Lifespan Years Avg"] = (df["Lifespan Years Min"] + df["Lifespan Years Max"]) / 2
df["Cost per kWh (₹)"] = df["Replacement Cost Avg (₹)"] / df["Battery Capacity (kWh)"]

# Save to Excel
with pd.ExcelWriter("EV_Battery_Analysis.xlsx", engine="openpyxl") as writer:
    # Main data sheet
    df.to_excel(writer, sheet_name="Battery Data", index=False)
    
    # Summary statistics sheet
    summary = df.describe().loc[["mean", "min", "max"]]
    summary.to_excel(writer, sheet_name="Summary Statistics")
    
    # Get workbook and worksheet objects for formatting
    workbook = writer.book
    worksheet = writer.sheets["Battery Data"]
    
    # Create styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Auto-adjust column widths
    for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(column))
        col_letter = get_column_letter(df.columns.get_loc(column) + 1)
        worksheet.column_dimensions[col_letter].width = column_length + 2

print("Excel file 'EV_Battery_Analysis.xlsx' created successfully!")